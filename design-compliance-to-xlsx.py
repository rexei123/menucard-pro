#!/usr/bin/env python3
# design-compliance-to-xlsx.py
# Wandelt tests/design-compliance/report.json in eine Excel-Mappe mit mehreren Sheets.

import json, re, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
REPORT = os.path.join(HERE, 'report.json')
OUT    = os.path.join(HERE, 'DESIGN-COMPLIANCE-REPORT.xlsx')

if len(sys.argv) >= 2: REPORT = sys.argv[1]
if len(sys.argv) >= 3: OUT    = sys.argv[2]

with open(REPORT, 'r', encoding='utf-8') as f:
    data = json.load(f)

pages = data.get('pages', [])

# ---------- Styles ----------
BOLD   = Font(name='Arial', bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', start_color='1A1A1A')
OK_FILL  = PatternFill('solid', start_color='DCFCE7')
FAIL_FILL= PatternFill('solid', start_color='FEE2E2')
WARN_FILL= PatternFill('solid', start_color='FEF3C7')
thin   = Side(border_style='thin', color='E5E7EB')
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

def header(ws, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = BOLD
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER
    ws.freeze_panes = 'A2'

def autosize(ws, max_width=60):
    for col in ws.columns:
        L = min(max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2, max_width)
        ws.column_dimensions[col[0].column_letter].width = max(L, 10)

wb = Workbook()

# ===== Sheet 1: Übersicht =====
ws = wb.active
ws.title = 'Übersicht'
header(ws, ['Status','Kind','Label','Viewport','URL','HTTP','ms','Token-Fehler','Font-Check','Emojis','Material-Icons','Fehler'])
row = 2
summary = {'admin':{'pass':0,'fail':0},'guest-menu':{'pass':0,'fail':0},'guest-item':{'pass':0,'fail':0}}

for p in pages:
    tokens = p.get('tokens') or {}
    token_fail = sum(1 for v in tokens.values() if not v.get('pass'))
    token_total = len(tokens)
    fonts = p.get('fonts') or {}
    icons = p.get('icons') or {}
    body_font = (fonts.get('bodyFont') or '')
    head_font = (fonts.get('headFont') or '')
    kind = p.get('kind','')
    # Erwartete Fonts:
    if kind == 'admin':
        font_ok = bool(re.search(r'Roboto', body_font, re.I)) and bool(re.search(r'Roboto', head_font, re.I))
        font_note = f"body={body_font[:40]} | head={head_font[:40]}"
    else:
        tpl = (p.get('extra') or {}).get('templateKey','').lower()
        expect = {
            'elegant':  (r'Playfair Display', r'Inter'),
            'modern':   (r'Montserrat',       r'Montserrat'),
            'classic':  (r'Playfair Display', r'(Playfair|Inter)'),
            'minimal':  (r'(Grotesk|Inter|Roboto)', r'(Grotesk|Inter|Roboto)'),
        }.get(tpl, (r'.', r'.'))
        font_ok = bool(re.search(expect[0], head_font, re.I)) and bool(re.search(expect[1], body_font, re.I))
        font_note = f"[{tpl}] body={body_font[:30]} | head={head_font[:30]}"

    emoji_count = icons.get('emojiCount', 0)
    mat_count   = icons.get('materialCount', 0)
    # Gesamtstatus
    status = 'PASS' if (token_fail == 0 and font_ok and emoji_count == 0) else 'FAIL'
    bucket = 'guest-menu' if kind=='guest-menu' else ('guest-item' if kind=='guest-item' else 'admin')
    summary.setdefault(bucket, {'pass':0,'fail':0})[('pass' if status=='PASS' else 'fail')] += 1

    cells = [
        status, kind, p.get('label',''),
        f"{p['viewport']['width']}x{p['viewport']['height']}",
        p.get('url',''), p.get('status',0), p.get('ms',0),
        f"{token_fail}/{token_total}", font_note, emoji_count, mat_count,
        (p.get('errMsg') or '')[:120],
    ]
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BORDER
        if i == 1:
            c.fill = OK_FILL if status=='PASS' else FAIL_FILL
            c.font = Font(bold=True, color='16A34A' if status=='PASS' else 'B91C1C')
    row += 1

autosize(ws, 60)

# ===== Sheet 2: Token-Details =====
ws2 = wb.create_sheet('Tokens')
header(ws2, ['Seite','Viewport','Token','Soll','Ist','Soll-kanonisch','Ist-kanonisch','Status'])
r = 2
for p in pages:
    for k, v in (p.get('tokens') or {}).items():
        ws2.append([p.get('label',''), f"{p['viewport']['width']}x{p['viewport']['height']}",
                    k, v.get('expected',''), v.get('actual',''),
                    v.get('canonExpected',''), v.get('canonActual',''),
                    'PASS' if v.get('pass') else 'FAIL'])
        c = ws2.cell(row=r, column=8)
        c.fill = OK_FILL if v.get('pass') else FAIL_FILL
        c.font = Font(bold=True, color='16A34A' if v.get('pass') else 'B91C1C')
        r += 1
autosize(ws2, 60)

# ===== Sheet 3: Fonts =====
ws3 = wb.create_sheet('Fonts')
header(ws3, ['Seite','Kind','Template','Body-Font','Heading-Font','Material-Symbols geladen'])
for p in pages:
    f = p.get('fonts') or {}
    ws3.append([p.get('label',''), p.get('kind',''),
                (p.get('extra') or {}).get('templateKey',''),
                f.get('bodyFont',''), f.get('headFont',''),
                'JA' if f.get('hasMaterial') else 'NEIN'])
autosize(ws3, 50)

# ===== Sheet 4: Icons =====
ws4 = wb.create_sheet('Icons')
header(ws4, ['Seite','Kind','Emoji-Count','Unique-Emojis','Material-Count','Fehlende Core-Icons','Beispiel-Icons'])
for p in pages:
    i = p.get('icons') or {}
    ws4.append([p.get('label',''), p.get('kind',''),
                i.get('emojiCount',0),
                ' '.join(i.get('uniqueEmojis',[])[:20]),
                i.get('materialCount',0),
                ', '.join(i.get('missingCore',[])[:20]),
                ', '.join(i.get('presentIcons',[])[:15])])
autosize(ws4, 60)

# ===== Sheet 5: Layout =====
ws5 = wb.create_sheet('Layout')
header(ws5, ['Seite','Kind','Viewport','Sidebar-Breite','Sidebar-Sel','Header-Höhe','Header-Sel','List-Panel-Breite','List-Sel','Soll Sidebar','Soll Header','Soll List'])
for p in pages:
    l = p.get('layout') or {}
    vp = f"{p['viewport']['width']}x{p['viewport']['height']}"
    ws5.append([p.get('label',''), p.get('kind',''), vp,
                l.get('sidebarWidth'), l.get('sidebarSel',''),
                l.get('headerHeight'), l.get('headerSel',''),
                l.get('listWidth'), l.get('listSel',''),
                200, 56, 400])
autosize(ws5, 40)

# ===== Sheet 6: Zusammenfassung =====
ws6 = wb.create_sheet('Zusammenfassung')
ws6['A1'] = 'Design-Compliance-Report — MenuCard Pro gegen Design-Strategie 2.0'
ws6['A1'].font = Font(name='Arial', bold=True, size=14)
ws6['A2'] = f"Erzeugt: {data.get('generatedAt','')}"
ws6['A3'] = f"Basis:   {data.get('base','')}"
ws6['A5'] = 'Bereich';       ws6['B5'] = 'PASS'; ws6['C5'] = 'FAIL'
for c in ['A5','B5','C5']:
    ws6[c].font = BOLD; ws6[c].fill = HDR_FILL
r = 6
for k, v in summary.items():
    ws6.cell(row=r, column=1, value=k)
    ws6.cell(row=r, column=2, value=v['pass']).fill = OK_FILL
    ws6.cell(row=r, column=3, value=v['fail']).fill = FAIL_FILL
    r += 1
ws6.cell(row=r, column=1, value='GESAMT').font = Font(bold=True)
ws6.cell(row=r, column=2, value=f"=SUM(B6:B{r-1})").font = Font(bold=True)
ws6.cell(row=r, column=3, value=f"=SUM(C6:C{r-1})").font = Font(bold=True)

ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 10
ws6.column_dimensions['C'].width = 10

# Sheet-Reihenfolge: Zusammenfassung vorne
wb.move_sheet('Zusammenfassung', offset=-5)

wb.save(OUT)
print(f"OK — {OUT}")
print(f"Seiten: {len(pages)}  |  Summary: {summary}")
